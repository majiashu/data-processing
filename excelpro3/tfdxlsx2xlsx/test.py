# -*- encoding: utf-8 -*-
"""
test.py
Created on 2018/8/9 14:36
Copyright (c) 2018/8/9, 
@author: 马家树(majstx@163.com)
"""
import openpyxl

wb = openpyxl.load_workbook("C:/Users/meridian/Desktop/test/test.xlsx")
ws = wb.get_active_sheet()
val = ws.cell(row=1, column=1).value

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.cell(row=1, column=1).value = val