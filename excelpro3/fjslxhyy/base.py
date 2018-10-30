# -*- encoding: utf-8 -*-
"""
base.py
Created on 2018/10/26 11:44
@author: 马家树(majstx@163.com)
"""
import openpyxl
from utils.myutil import *


def jg2xlsx():
    row_num = 2
    for r in range(2, max_row_num + 1):
        tjbh = ws.cell(row=r, column=1).value
        lasttjbh = ws.cell(row=r - 1, column=1).value
        if tjbh != lasttjbh:
            row_num = row_num + 1
        xmmc = ws.cell(row=r, column=2).value
        jg = ws.cell(row=r, column=3).value
        jg = str_formatxm(jg)
        dw = ws.cell(row=r, column=4).value
        fw = ws.cell(row=r, column=5).value
        fwdw = str(fw) + '[' + str(dw) + ']'

        # 数值分类  判断该结果是正常？偏高？偏低？
        if is_number(jg):
            try:
                fl_value = compare(jg, fw)
            except:
                fl_value = jg
        else:
            fl_value = jg

        # 如果指标在字典中就取出该列，不在添加表头
        if xmmc in dict_allindex:
            col_num = dict_allindex[xmmc]
        else:
            col_num = len(dict_allindex) + 1
            dict_allindex[xmmc] = col_num  # 添加进字典
            cell_bt = number2char(col_num) + '1'
            sheet[cell_bt] = xmmc
            sheet1[cell_bt] = xmmc
            sheet2[cell_bt] = xmmc

        # 定位此时单元格的位置
        cell_num = number2char(col_num) + str(row_num)

        # 写结果
        sheet[cell_num] = jg
        sheet1[cell_num] = fl_value
        sheet2[cell_num] = fwdw

        # 写体检编号
        cell_num = 'A' + str(row_num)
        sheet[cell_num] = tjbh
        sheet1[cell_num] = tjbh
        sheet2[cell_num] = tjbh


if __name__ == '__main__':
    inpath = "C:/Users/meridian/Desktop/福建省立协和医院.xlsx"
    outpath = "C:/Users/meridian/Desktop/out"
    hospname = "福建省立协和医院"

    # 读取
    wb = openpyxl.load_workbook(inpath)
    ws = wb.get_active_sheet()
    max_row_num = ws.max_row

    # 写
    workbook = openpyxl.Workbook()
    workbook1 = openpyxl.Workbook()
    workbook2 = openpyxl.Workbook()
    sheet = workbook.active  # 原始表
    sheet1 = workbook1.active  # 分类表
    sheet2 = workbook2.active  # 范围表

    # 结果
    pebt_list = ['体检编号']
    sheet.append(pebt_list)
    sheet1.append(pebt_list)
    sheet2.append(pebt_list)
    dict_allindex = dict(zip(pebt_list, range(1, len(pebt_list) + 1)))
    jg2xlsx()

    printlog('save source...')
    workbook.save(outpath + '/' + hospname + "_原始.xlsx")
    printlog('save fl...')
    workbook1.save(outpath + '/' + hospname + "_分类.xlsx")
    printlog('save fwdw...')
    workbook2.save(outpath + '/' + hospname + "_范围.xlsx")
    printlog('save over!')
