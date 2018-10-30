# -*- encoding: utf-8 -*-
"""
extract_source.py
Created on 2018/10/9 11:55
@author: 马家树(majstx@163.com)
"""

import openpyxl
import pymysql
import xlsxwriter


# 读
wb = openpyxl.load_workbook('C:/Users/meridian/Desktop/轻松筹相关指标解读.xlsx')
sheet = wb.get_sheet_by_name('轻松筹指标解读')

# 写入
workbook1 = xlsxwriter.Workbook(filename=u'C:\\Users\\meridian\\Desktop\\test0\\out.xlsx')
workbook1.use_zip64()
sheet1 = workbook1.add_worksheet()

conn = pymysql.connect(host='10.1.1.102', port=3306, user='query', password='123456',
                       database='checkup_library', charset='utf8mb4')
cur = conn.cursor()
sql = '''SELECT code_content FROM `code_content` WHERE index_code = '''

row_max = sheet.max_row
row_max = row_max + 1
for r in range(3, row_max):
    key = sheet.cell(row=r, column=2).value
    cur.execute(sql + "'" + key + "'")
    source_list = cur.fetchall()
    try:
        source1 = source_list[0][0]
    except:
        source1 = ''
    try:
        source2 = source_list[1][0]
    except:
        source2 = ''
    try:
        source3 = source_list[2][0]
    except:
        source3 = ''
    print(source1, source2, source3)

    print(source_list)
    sheet1.write(r, 0, key)
    sheet1.write(r, 1, source1)
    sheet1.write(r, 2, source2)
    sheet1.write(r, 3, source3)
workbook1.close()


