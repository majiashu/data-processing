# -*- encoding: utf-8 -*-
"""
wordcount.py
Created on 2018/9/3 14:54
@author: 马家树(majstx@163.com)
"""

import openpyxl

wb = openpyxl.load_workbook('C:/Users/meridian/Desktop/新建文件夹/test.xlsx')
ws = wb.get_active_sheet()
col_num = ws.max_column
row_num = ws.max_row
val_dic = {}
for c in range(1, col_num+1):
    print(c)
    for r in range(1, row_num+1):
        val = ws.cell(row=r, column=c).value
        val = str(val)
        val_list = val.split('；')
        print(val_list)
        for v in val_list:
            val_dic.setdefault(v, 0)
            val_dic[v] += 1
print(val_dic)