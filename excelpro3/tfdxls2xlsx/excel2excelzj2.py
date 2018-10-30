# -*- encoding: utf-8 -*-
"""
excel2excel5.py
Created on 2018/1/2 17:17
Copyright (c) 2018/1/2, 
@author: 小马同学(majstx@163.com)
"""

import xlrd
import openpyxl
import os

"""
提取天方达.xls文件的综述和建议,该综述建议没有结束标志 体检结论：
"""


def xls2xlsx():
    # 人的信息
    tjbh = file.split('_')[0]

    # 项目的信息
    for i in range(0, nrows):
        flag = sheetd.cell(i, 0).value
        if flag == '综  述：':
            zsi = i
        if flag == '建  议：':
            jyi = i

    if not 'jyi' in locals().keys():  # 判断是否有建议，也就是看jyi是否定义
        # 综述
        zsvaluez = ''
        for i in range(zsi, jyi):
            zsvalue = sheetd.cell(i, 1).value
            zsvaluez = zsvaluez + zsvalue
        # 建议
        jyvaluez = ''
    else:
        # 综述
        zsvaluez = ''
        for i in range(zsi, jyi):
            zsvalue = sheetd.cell(i, 1).value
            zsvaluez = zsvaluez + zsvalue
        # 建议
        jyvaluez = ''
        for i in range(jyi, nrows):
            jyvalue = sheetd.cell(i, 1).value
            jyvaluez = jyvaluez + jyvalue

    sheet.cell(row=row_num, column=1).value = tjbh
    sheet.cell(row=row_num, column=2).value = zsvaluez
    sheet.cell(row=row_num, column=3).value = jyvaluez


if __name__ == '__main__':
    workbook = openpyxl.Workbook()
    sheet = workbook.active  # 原始表
    pebt_list = ['体检编号', '综述', '建议']
    sheet.append(pebt_list)
    # dict_allindex = dict(zip(pebt_list, range(1, len(pebt_list) + 1)))

    path = u"E:/原始数据/xls格式的数据/松滋人民医院/总/"  # 数据源路径
    files = os.listdir(path)
    row_num = 2
    for file in files:

        # print u"现在处理的的是第%d个人的数据" % row_num
        try:
            wbd = xlrd.open_workbook(u"E:/原始数据/xls格式的数据/松滋人民医院/总/%s" % file)
        except:
            print("{0}该文件格式无效".format(file))
            continue
        sheetd = wbd.sheets()[0]
        nrows = sheetd.nrows
        ncols = sheetd.ncols
        xls2xlsx()

        row_num += 1

    print('save zj...')
    workbook.save(u"C:/Users/meridian/Desktop/处理后的数据(3.6以后)/松滋人民医院out/松滋人民医院__综述建议.xlsx")
    print('save over')
