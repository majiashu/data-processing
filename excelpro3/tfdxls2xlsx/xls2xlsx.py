# -*- encoding: utf-8 -*-
"""
xls2xlsx.py
Created on 2018/8/28 13:30
@author: 马家树(majstx@163.com)
"""

# 提取天方达xls格式的体检报告单，包括原始表 分类表 范围表 小结表 综述建议表
# 原文件要有 “体检结论：” 结束标志

from utils.myutil import *
import xlrd
import openpyxl
import os


def jg2xlsx():
    # 人的信息
    peinfo = []
    tjbh = file.split('_')[0]
    khbh = file.split('-')[0]
    xm = file.split('_')[1].split('.')[0]
    peinfo.append(tjbh)
    peinfo.append(khbh)
    peinfo.append(xm)

    peinfo1 = sheetd.cell(1, 0).value
    pos = peinfo1.find('性别：')
    peinfo1 = peinfo1[pos:]  # 性别：男  年龄：65
    list1 = peinfo1.split()
    peinfo2 = sheetd.cell(2, 0).value
    list2 = peinfo2.split()
    list = list1 + list2
    for i in range(0, len(list)):
        p1 = list[i]
        pos = p1.find('：')
        p1 = p1[pos + 1:]
        peinfo.append(p1)

    for col_num in range(1, len(peinfo) + 1):
        cell_num = number2char(col_num) + str(row_num)
        col_num -= 1
        value = peinfo[col_num]
        sheet[cell_num] = value  # 添加个人信息
        sheet1[cell_num] = value
        sheet2[cell_num] = value

    # 项目的信息

    for i in range(0, nrows):
        flag = sheetd.cell(i, 0).value
        if flag == '项目名称':
            ksvalue = sheetd.cell(i - 2, 0).value
            if ksvalue == '小  结：':  #
                ksvalue = lastksvalue
            else:
                ksvalue = sheetd.cell(i - 2, 0).value
            dxvalue = sheetd.cell(i - 1, 0).value
            # 小项及体检结果
            for r in range(i + 1, nrows):  # 行
                xxvalue = sheetd.cell(r, 0).value
                if xxvalue == '小  结：':
                    break
                index = ksvalue + '_' + dxvalue + '_' + xxvalue
                jgvalue = sheetd.cell(r, 1).value
                jgvalue = str_formatxm(jgvalue)
                dwvalue = sheetd.cell(r, 2).value
                fwvalue = sheetd.cell(r, 3).value
                fwdwvalue = fwvalue + '[' + dwvalue + ']'
                if fwdwvalue == '[]':
                    fwdwvalue = ''
                elif fwdwvalue == '-[]':
                    fwdwvalue = ''

                # 数值分类  判断该结果是正常？偏高？偏低？
                if is_number(jgvalue):
                    try:
                        fl_value = compare(jgvalue, fwvalue)
                    except:
                        fl_value = jgvalue
                else:
                    fl_value = jgvalue

                # 如果指标在字典中就取出该列，不在添加表头
                if index in dict_allindex:
                    col_num = dict_allindex[index]
                else:
                    col_num = len(dict_allindex) + 1
                    dict_allindex[index] = col_num  # 添加进字典
                    cell_bt = number2char(col_num) + '1'
                    sheet[cell_bt] = index
                    sheet1[cell_bt] = index
                    sheet2[cell_bt] = index
                # 定位此时单元格的位置
                cell_num = number2char(col_num) + str(row_num)

                # 对结果按照预定的规则格式化

                sheet[cell_num] = jgvalue
                sheet1[cell_num] = fl_value
                sheet2[cell_num] = fwdwvalue

            lastksvalue = ksvalue


def xj2xlsx():
    # 人的信息
    peinfo = []
    tjbh = file.split('_')[0]
    peinfo.append(tjbh)

    for col_num in range(1, len(peinfo) + 1):
        cell_num = number2char(col_num) + str(row_num)
        col_num -= 1
        value = peinfo[col_num]
        sheet3[cell_num] = value  # 添加个人信息
        # sheet2[cell_num] = value

    # 项目的信息

    for i in range(0, nrows):
        flag = sheetd.cell(i, 0).value
        if flag == '项目名称':
            ksvalue = sheetd.cell(i - 2, 0).value
            if ksvalue == '小  结：':
                ksvalue = lastksvalue
            else:
                ksvalue = sheetd.cell(i - 2, 0).value
            dxvalue = sheetd.cell(i - 1, 0).value
            # 小项及体检结果
            for r in range(i + 1, nrows):  # 行
                xxvalue = sheetd.cell(r, 0).value
                if xxvalue == '小  结：':
                    xxvalue = '小结'
                    index = ksvalue + '_' + dxvalue + '_' + xxvalue
                    xjvalue = sheetd.cell(r, 1).value
                    xjvalue = str_formatxm(xjvalue)

                    # 如果指标在字典中就取出该列，不在添加表头
                    if index in dict_allindex:
                        col_num = dict_allindex[index]
                    else:
                        col_num = len(dict_allindex) + 1
                        dict_allindex[index] = col_num  # 添加进字典
                        cell_bt = number2char(col_num) + '1'
                        sheet3[cell_bt] = index
                    # 定位此时单元格的位置
                    cell_num = number2char(col_num) + str(row_num)
                    try:
                        sheet3[cell_num] = xjvalue
                    except:
                        print('设置异常')
                    break

            lastksvalue = ksvalue


def zsjy2xlsx():
    # 人的信息
    tjbh = file.split('_')[0]

    # 项目的信息
    for i in range(0, nrows):
        flag = sheetd.cell(i, 0).value
        if flag == '综  述：':
            zsi = i
        if flag == '建  议：':
            jyi = i
        if flag == '体检结论：':
            tjjli = i

    if not 'jyi' in locals().keys():  # 判断是否有建议，也就是看jyi是否定义
        # 综述
        zsvaluez = ''
        for i in range(zsi, tjjli):
            zsvalue = sheetd.cell(i, 1).value
            zsvaluez = zsvaluez + zsvalue
        # 建议
        jyvaluez = ''
    else:
        # 综述
        zsvaluez = ''
        for i in range(zsi, jyi):
            zsvalue = sheetd.cell(i, 1).value
            zsvaluez = zsvaluez + zsvalue
        # 建议
        jyvaluez = ''
        for i in range(jyi, tjjli):
            jyvalue = sheetd.cell(i, 1).value
            jyvaluez = jyvaluez + jyvalue
    try:
        sheet4.cell(row=row_num, column=1).value = tjbh
        sheet4.cell(row=row_num, column=2).value = zsvaluez
        sheet4.cell(row=row_num, column=3).value = jyvaluez
    except:
        print('设置异常')


if __name__ == '__main__':

    inpath = "C:/Users/meridian/Desktop/in"
    outpath = "C:/Users/meridian/Desktop/out"
    hospname = "陕西省人民医院"

    workbook = openpyxl.Workbook()
    workbook1 = openpyxl.Workbook()
    workbook2 = openpyxl.Workbook()
    workbook3 = openpyxl.Workbook()
    workbook4 = openpyxl.Workbook()
    sheet = workbook.active  # 原始表
    sheet1 = workbook1.active  # 分类表
    sheet2 = workbook2.active  # 范围表
    sheet3 = workbook3.active  # 小结表
    sheet4 = workbook4.active  # 综述建议

    # 结果
    pebt_list = ['体检编号', '客户编号', '姓名', '性别', '年龄', '单位名称', '体检日期']
    sheet.append(pebt_list)
    sheet1.append(pebt_list)
    sheet2.append(pebt_list)
    dict_allindex = dict(zip(pebt_list, range(1, len(pebt_list) + 1)))

    # 小结表
    pebt_listxj = ['体检编号']
    sheet3.append(pebt_listxj)
    dict_allindexxj = dict(zip(pebt_listxj, range(1, len(pebt_listxj) + 1)))
    files = os.listdir(inpath)

    # 综述建议
    pebt_listzj = ['体检编号', '综述', '建议']
    sheet4.append(pebt_listzj)

    row_num = 2
    for file in files:
        # print   file
        if row_num % 500 == 0:
            printlog("现在处理的的是第{0}个人的数据".format(row_num))
        try:
            wbd = xlrd.open_workbook(inpath + '/' + file)
        except:
            printlog("{0}该文件格式无效".format(file))
            continue
        sheetd = wbd.sheets()[0]
        nrows = sheetd.nrows
        ncols = sheetd.ncols
        jg2xlsx()
        xj2xlsx()
        zsjy2xlsx()
        row_num += 1

    printlog('save source...')
    workbook.save(outpath + '/' + hospname + "_原始.xlsx")
    printlog('save fl...')
    workbook1.save(outpath + '/' + hospname + "_分类.xlsx")
    printlog('save fwdw...')
    workbook2.save(outpath + '/' + hospname + "_范围.xlsx")
    printlog('save xj...')
    workbook3.save(outpath + '/' + hospname + "_小结.xlsx")
    printlog('save zsjy...')
    workbook4.save(outpath + '/' + hospname + "_综述建议.xlsx")
    printlog('Save completion')
