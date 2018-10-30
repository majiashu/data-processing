# -*- encoding: utf-8 -*-
"""
cleanexcel.py
Created on 2018/8/21 13:36
Copyright (c) 2018/8/21, 
@author: 马家树(majstx@163.com)
"""
import openpyxl
from utils import myutil
import time


def main():
    # 读
    wbr = openpyxl.load_workbook('C:/Users/meridian/Desktop/襄阳航空工业医院/襄阳航空工业/exportdata_source.xlsx')
    shr = wbr.get_active_sheet()
    nrows = shr.max_row
    ncols = shr.max_column

    # 写
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for l in range(1, ncols+1):
        print(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()), '当前处理的是第%s列' % l)
        for r in range(1, nrows+1):
            val = shr.cell(row=r, column=l).value
            if l > 15:
                val = myutil.str_formatxm(val)
            sheet.cell(row=r, column=l).value = val
    print(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()), '开始保存！')
    workbook.save('C:/Users/meridian/Desktop/襄阳航空工业医院（新）/exportdata_source.xlsx')
    print(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()), '保存成功！')


if __name__ == '__main__':
    main()
