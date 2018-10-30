# -*- encoding: utf-8 -*-
"""
excel_split.py
Created on 2018/4/26 10:27
Copyright (c) 2018/4/26, 
@author: 马家树(majstx@163.com)
"""

from time import time
import openpyxl
import xlsxwriter

"""
本模块就是按照字典对应拆分文件
"""


def readdic():
    """
    读取指标对应的分类字典
    :return: dic_classify
    """
    start = time()
    wb = openpyxl.load_workbook(u'C:\\Users\\meridian\\Desktop\\test0\\exportdata_classify-变量分类.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    print ('文件读取成功. Cost {:.3f}s'.format(time() - start))
    row_num = sheet.max_row
    # col_num = sheet.max_column
    dic_classify = {}
    for r in range(1, row_num+1):
        key_classify = sheet.cell(row=r, column=1).value
        value_classify = sheet.cell(row=r, column=2).value
        dic_classify[key_classify] = value_classify
    return dic_classify


def splitfile():
    """
    将文件拆分主逻辑
    :return:
    """
    dic_classify = readdic()
    print('读取文件')
    start = time()
    wb = openpyxl.load_workbook(u'C:\\Users\\meridian\\Desktop\\范围\\数值类\\exportdata_ranges2（删除了空指标）.xlsx', read_only=True)
    print ('文件读取成功. Cost {:.3f}s'.format(time() - start))
    rows = wb.active.rows

    # 写入
    workbook1 = xlsxwriter.Workbook(filename=u'C:\\Users\\meridian\\Desktop\\范围\\数值类\\exportdata_ranges2（删除了空指标）(数值类).xlsx')
    workbook1.use_zip64()
    sheet1 = workbook1.add_worksheet()

    indexs_dict_r = {}
    # indexs_dict_w = {}
    for row_num, row in enumerate(rows):
        for col_num, cell in enumerate(row):
            val = cell.value.strip() if cell.value is not None else ''
            if row_num == 0:
                indexs_dict_r[col_num] = val
                if dic_classify[indexs_dict_r[col_num]] == u'数值类' or dic_classify[indexs_dict_r[col_num]] == u'基本信息':
                    sheet1.write(row_num, col_num, val)
            else:
                if dic_classify[indexs_dict_r[col_num]] == u'数值类' or dic_classify[indexs_dict_r[col_num]] == u'基本信息':
                    sheet1.write(row_num, col_num, val)
                else:
                    continue


if __name__ == '__main__':
    splitfile()
